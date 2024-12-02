import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver import ActionChains
import time
from openpyxl import load_workbook

workbook = load_workbook(r'C:\Users\senthamarai.kannan\OneDrive - OneWorkplace\Desktop\SELENIUM\testdata.xlsx')
sheet = workbook.active

driver = webdriver.Firefox()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    time.sleep(5)

    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Username']").send_keys(username)
    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Password']").send_keys(password)

    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".oxd-userdropdown-name").click()
    time.sleep(2)
    driver.find_element(By.LINK_TEXT, "Logout").click()
    time.sleep(2)

driver.quit()
